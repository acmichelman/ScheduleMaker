import sqlite3
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill 

from .. import clear_screen, main_menu
from ..db import DB_PATH, ensure_db_dir
#DB_PATH = Path(__file__).resolve().parents[1] / "DatabaseFold" / "TOHLifeguardDB"

"""
Civic            Middle               2 Chair           Organize the schedule so Senior Lieutenant-> Lieutenant-> Senior Guard-> Lifeguard
Sr Lieutenant    J.Doe (Mon, Tue)     J.Rock (Wes, Thur)
Lieutenant       S.Chase (Wed, Thur)  D.Will (Fri, Sun)
Sr Guard         ect                  ect
Lifeguard

Tired of looking up links
https://www.geeksforgeeks.org/python/working-with-excel-spreadsheets-in-python/
https://openpyxl.readthedocs.io/en/stable/
https://www.youtube.com/watch?v=0AlQtxFqv54
"""
#   On our schedule we have lieutenants names highlighted in yellow and senior guards highlighted in blue. Makes it easier for young guards to know whos in charge
YELLOW_FILL = PatternFill(start_color="f7cb52", end_color="f7cb52", fill_type="solid")
BLUE_FILL   = PatternFill(start_color="54aaeb", end_color="54aaeb", fill_type="solid")

#   We want the schedule to be ordered by highest rank to lowest
RANK_ORDER = {
    "senior lieutenant": 0,
    "lieutenant": 1,
    "senior guard": 2,
    "lifeguard": 3,
}

def get_past_full_schedule(schedule_period: str):
    with sqlite3.connect(DB_PATH) as con:
        cur = con.cursor()
        cur.execute("""SELECT EmpDaysOff, BeachNameText, FirstLastNameText, EmpRank
                    FROM Schedules
                    WHERE SchedulePeriod = ?;
                    """, (schedule_period,))
        rows = cur.fetchall()
        
        schedules_by_row_id: dict[int, dict] = {}

        for i, (days_off, beach_name, full_name, emp_rank) in enumerate(rows, start=1):
            schedules_by_row_id[i] = {
                "RowID": i,
                "SchedulePeriod": schedule_period,
                "EmpDaysOff": days_off,
                "BeachNameText": beach_name,
                "FirstLastNameText": full_name,
                "EmpRank": emp_rank,
            }

    return schedules_by_row_id

#   Looks at Schedules DB and gets all unique 2 week periods as options to export
def get_past_schedulePeriod():
    with sqlite3.connect(DB_PATH) as con:
        cur = con.cursor()
        cur.execute("""SELECT DISTINCT SchedulePeriod
                    FROM Schedules
                    """)
        rows = cur.fetchall()
        return rows
    
#   Export to excel logic
def export_to_excel(period_option: str):
    rows = get_past_full_schedule(period_option)
    
    #   should get a hashmap of BeachID so we can order columns by BeachID
    with sqlite3.connect(DB_PATH) as con:
        cur = con.cursor()
        cur.execute("SELECT BeachID, BeachName FROM Beaches;")
        beach_rows = cur.fetchall()

    #-----Bucket the schedule
    #   Bucket the schedules under a beach id (ex. Civic-> name1, name2, name3 Middle-> name4, name5)
    beach_id_by_name: dict[str, int] = {} # For sorting colum id
    beach_name_by_id: dict[int, str] = {} # For writting the header text
    for beach_id, beach_name in beach_rows: #   TODO: sort with just one dict maybe. Right now it makes sense but redundent?
        if beach_name is None:  #TODO: We can probally just remove all of this
            continue
        bn = str(beach_name)
        beach_id_by_name[bn.lower()] = int(beach_id) #  We have the first letter captalized so this should make fuck ups less likely
        beach_name_by_id[int(beach_id)] = bn #  Put the uncaplized back in its box (back in the dict)

    #-----Sorting guards by rank
    schedule_by_beach: dict[int, list[tuple[int, str, str]]] = {}

    for _, r in rows.items(): # We need a way to organize the data to match what usual schedule look like
        beach_text = str(r.get("BeachNameText", ""))
        name_text  = str(r.get("FirstLastNameText", ""))
        days_off   = str(r.get("EmpDaysOff", ""))
        rank_text  = str(r.get("EmpRank", ""))

        if not beach_text:
            continue

        #   Matchs Beaches table so we get BeachID
        beach_id = beach_id_by_name.get(beach_text.lower())
        if beach_id is None:
            continue

        #   Sort the data so ranks are hierarchical 
        rank_sort = RANK_ORDER.get(rank_text, 999)
        if days_off:
            cell_text = f"{name_text} ({days_off})"
        else:
            cell_text = name_text
        name_sort = name_text.lower()

        schedule_by_beach.setdefault(beach_id, []).append((rank_sort, name_sort, cell_text, rank_text)) #  Data should be organised here

    #-----Sort by beach id
    #   We want the beach Id to be placed in excel following the order it was inserted or if no other was inserted following what the beach looks like
    beach_ids = sorted(schedule_by_beach.keys())
    for bid in beach_ids:
        schedule_by_beach[bid].sort(key=lambda t: (t[0], t[1]))  #  rank then name

    #-----Export to Excel
    #   nightmare nightmare nightmare. Im no longer having fun with this part
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    #   Header row (Will be our beach names)
    for col_idx, bid in enumerate(beach_ids, start=1):
        ws.cell(row=1, column=col_idx).value = beach_name_by_id.get(bid, f"Beach {bid}")

    #   Body rows (Will be our employees/ days off). Each Beach is a column with ranks in descending order
    max_len = max(len(schedule_by_beach[bid]) for bid in beach_ids)
    for i in range(max_len):
        row_num = 2 + i
        for col_idx, bid in enumerate(beach_ids, start=1):
            emps = schedule_by_beach[bid]
            if i >= len(emps):
                continue  # nothing in this column on this row

            #   Color the cell based on rank
            rank_sort, name_sort, cell_text, rank_text = emps[i]
            c = ws.cell(row=row_num, column=col_idx)
            c.value = cell_text

            c.value = cell_text
            if rank_text in ("senior lieutenant", "lieutenant"):
                c.fill = YELLOW_FILL
            elif rank_text == "senior guard":
                c.fill = BLUE_FILL

    #   We do a bit of formatting
    ws.freeze_panes = "A2"
    for col_idx in range(1, len(beach_ids) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 30

    #   Save to Downloads. Like your computers download folder. I mean I hope
    downloads = Path.home() / "Downloads"
    out_dir = downloads if downloads.exists() else Path.cwd()

    clean_file_name = period_option.replace("/", "-").replace(" ", "") #    Make sure we strip all special characters in file name or we have Alexander's PC all over again
    out_path = out_dir / f"Schedule_{clean_file_name}.xlsx"
    wb.save(out_path)   #   please work this time

    print(f"Exported to: {out_path}")

def menu_options():
    print("\nMenu Options")
    print("Export Employees (Export)[1] \nBack[2]")

def export_emp_to_excel_menu():

    clear_screen.clear_screen()
    print("Welcome to the export employee page.\n")
    menu_options()

    running = True
    
    while running:
        ans = input("> ").strip()

        if ans.lower() == "export" or ans == "1":
            rows = get_past_schedulePeriod()#   want it to look like 1) 5/25/2026 - 06/07/2026 ect
            for i, (schedulePeriod,) in enumerate(rows, start= 1): #    i = incremented number for eachschedule period, schedulePeriod is our two week date, rows is the DB data
                print(f"{i}) {schedulePeriod}")
            print("Please select a schedule you wish to export to Excel")
            export_ans = input("> ").strip()
            if not export_ans.isdigit():
                print("Invalid choice")
                menu_options()
                continue
            export_ans = int(export_ans)
            export_ans -= 1
            if 0 <= export_ans <= len(rows) - 1:
                export_ans = rows[export_ans][0] 
                export_to_excel(export_ans)
            else:
                print("Invalid choice")

        elif ans.lower() == "back" or ans == "2":
            running = False
            main_menu.main_menu()

        else:
            print("Please pick a different option")
        menu_options()