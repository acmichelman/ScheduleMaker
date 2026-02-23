import sqlite3
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from .. import clear_screen, main_menu

#DB_PATH = Path(__file__).resolve().parents[1] / "DatabaseFold" / "TOHLifeguardDB"
from ..db import DB_PATH, ensure_db_dir
EXCEL_DIR = Path(__file__).resolve().parent  #  ExcelFolder/

#   Will use this to prompt the user if correct file was loaded
def find_newest_excel_file():
    #   Get newest .xlsx file in ExcelFolder
    excel_files = sorted(EXCEL_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not excel_files:
        return -1
        
    EXCEL_PATH = excel_files[0]  #  newest file
    #   Name + last modified timestamp
    file_name = EXCEL_PATH.name
    modified_dt = datetime.fromtimestamp(EXCEL_PATH.stat().st_mtime)
    
    print(f"Loaded Excel file: {file_name} (last modified: {modified_dt:%m/%d/%Y %I:%M %p})")
    #   wb = Workbook: The Excel file itself, containing one or more worksheets
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    return ws

#   RANK lg, sg, lt, sl
def normalize_rank(value) -> str:
    ans = str(value).strip().lower()
    mapping = {
        "lg": "lifeguard",
        "lifeguard": "lifeguard",
        "sg": "senior guard",
        "senior guard": "senior guard",
        "lt": "lieutenant",
        "lieutenant": "lieutenant",
        "sl": "senior lieutenant",
        "senior lieutenant": "senior lieutenant",
    } 
    return mapping.get(ans, ans)

#   DATE MM/DD/YYYY
def normalize_date(value) -> str:
    #Should take one instance/ colum-row at a time
    """
    Accepts:
      - Excel dates (datetime)
      - strings like MM/DD/YYYY
      - blank/None -> "NA"
    """
    #   If the value is empty we can return NA 
    if value is None or str(value).strip() == "":
        return "NA"
    
    #   Converts to approate datetime
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    
    #   wasnt datetime so its text
    s = str(value).strip()

    if s.lower() in ("na", "n/a", "none"):
        return "NA"
    
    #   enforce MM/DD/YYYY
    datetime.strptime(s, "%m/%d/%Y")
    return s

#   EVAL SCORE 1,2,3,4,5 (default is 4)
def normalize_eval(value) -> str:
    if value is None or str(value).strip() == "":
        return 4

    n = int(value)
    if n < 1 or n > 5:
        return -1 # MAKE SURE THIS IS SKIPPED
    return n

#   LOAD INTO DB
def load_excel_into_db(ws):

    #   Headers are our cells
    headers = [cell.value for cell in ws[1]]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    required = {"FirstName", "LastName", "EmployeeRank"}
    #   what required items are NOT present. Check if FirstName, LastName, EmployeeRank is there and for those that arent. We print those
    missing = required - set(headers)
    if missing:
        print("Excel is missing required columns:")
        for m in sorted(missing):
            print(f" - {m}")
        return #    End here if missing

    #   map column name -> index
    col = {name: headers.index(name) for name in headers if name != ""}

    #   For user to see what passed and what didnt
    inserted = 0
    skipped_blank_required = 0
    skipped_duplicates = 0
    skipped_bad_eval = 0
    errors = [] #   Will store all errors like ValueError, TypeError, etc

    with sqlite3.connect(DB_PATH) as con:
        cur = con.cursor()
        #   This has the program start at row 2 so it doesnt add FirstName, LastName, ect
        for excel_row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                #   Get the raw string from colums/row cells
                #   Grab raw FirstName cell
                if col["FirstName"] < len(row):
                    first_raw = row[col["FirstName"]]
                else:
                    first_raw = None
                #   Grab raw LastName cell
                if col["LastName"] < len(row):
                    last_raw  = row[col["LastName"]]
                else:
                    last_raw = None
                #   Grab raw EmployeeRank cell
                if col["EmployeeRank"] < len(row):
                    rank_raw  = row[col["EmployeeRank"]]
                else:
                    rank_raw = None

                #   Convert raw into clean string
                #   Convert FirstName to clean
                if first_raw is not None:
                    first = str(first_raw).strip()
                else:
                    first = ""
                #   Convert LastName to clean
                if last_raw is not None:
                    last  = str(last_raw).strip()
                else:
                    last = ""
                #   Convert EmployeeRank to clean
                if rank_raw is not None:
                    rank_text = str(rank_raw).strip()
                else:
                    rank_text =""

                #   SKIP IF BLANK and increment skipped_blank_required by 1, im just saying it over again
                if first == "" or last == "" or rank_text == "":
                    skipped_blank_required += 1
                    continue
                
                rank = normalize_rank(rank_text)

                if "DatePromoted" in col: 
                    if col["DatePromoted"] < len(row):
                        date_raw = row[col["DatePromoted"]]
                    else:
                        date_raw = None
                    date_promoted = normalize_date(date_raw)
                
                if "EvaluationScore" in col:
                    if col["EvaluationScore"] < len(row):
                        eval_raw = row[col["EvaluationScore"]]
                    else:
                        eval_raw = None
                    eval_score = normalize_eval(eval_raw)
                if eval_score == -1:
                    skipped_bad_eval += 1
                    continue
                
                #   Edge Case
                if rank == "lifeguard":
                    date_promoted = "NA"

                #   INSERT INTO OUR DATABASE !!!
                cur.execute("""INSERT INTO Employees
                            (FirstName, LastName, EmployeeRank, DatePromoted, EvaluationScore)
                            VALUES (?, ?, ?, ?, ?)
                            ON CONFLICT(FirstName, LastName) DO NOTHING;
                            """,(first, last, rank, date_promoted, eval_score))
                
                if cur.rowcount == 1:
                    inserted += 1
                else:
                    skipped_duplicates += 1
            #   Stored as e!!!!!!
            except Exception as e:
                errors.append((excel_row_num, str(e)))

        con.commit()

    #   Prints employees added/ errors that occured like skips and such.
    print("\nImport results:")
    print(f"Inserted: {inserted}")
    print(f"Skipped (blank First/Last/Rank): {skipped_blank_required}")
    print(f"Skipped (duplicates): {skipped_duplicates}")
    print(f"Skipped (bad eval score): {skipped_bad_eval}")

    if errors:
        print(f"\nErrors: {len(errors)} (showing up to 10)")
        for rnum, msg in errors[:10]:
            print(f"Row {rnum}: {msg}")



def menu_options():
    print("\nMenu Options")
    print("Import Employees (Import)[1] \nBack[2]")

def import_emp_from_excel_menu():
    #   Load employees from excel to speed up loading. I know we have a full employee file on work comp. Just need to format it a bit and find the god forsaken place its stored.
    clear_screen.clear_screen()
    print("Welcome to the import employee page\nPlease place any .xlsx (Excel) document into the ExcelFolder.")
    menu_options()

    running = True
    
    while running:
        ans = input("> ").strip()

        if ans.lower() == "Import" or ans == "1":
            print("Import Excel")
            ws = find_newest_excel_file() # Get newst excel file to be used
            if ws == -1:
                print("No .xlsx (Excel) file found in ExcelFolder.\nPlease Place a Excel file in the folder if you wish to load lifeguards.")
            else:
                print("Do you wish to use this Excel file Y/N")
                while True:
                    ans_correct_info = input().strip().lower()
                    if ans_correct_info == "yes" or ans_correct_info == "y":
                        load_excel_into_db(ws) #    insert the employees
                        break
                    elif ans_correct_info == "no" or ans_correct_info == "n":
                        ans = 0
                        break
                    else:
                        print("Please enter Yes or No (Y/N)")

        elif ans.lower() == "back" or ans == "2":
            running = False
            main_menu.main_menu()

        else:
            print("Please pick a different option")
        menu_options()